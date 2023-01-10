import copy
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
import glob
import os
from natsort import natsorted
# insert single image

# create a workbook and grab active worksheet
workbook = load_workbook(filename = 'template.xlsx')
# workbook = copy(rb)
worksheet = workbook.active 
# path to image
# img = Image('imagenes/60161.png')

# add single image
# ws.add_image(img, 'A1')

# # save workbook
# wb.save('image.xlsx')

# insert multiple images

# create a workbook and grab active worksheet
# workbook = Workbook()
# worksheet = workbook.active

# resize cells
for row in range(2,100):
    for col in range(7,9):
        worksheet.row_dimensions[row].height = 160
        col_letter = get_column_letter(col)
        worksheet.column_dimensions[col_letter].width = 40

# images list
images = []
for filename in natsorted(glob.glob('imagenes/*.png')):
    images.append(filename)

# insert images
for index, image in enumerate(images):
    worksheet.add_image(Image(image), anchor='H'+str(index+2))
    print(index,image)

# titles list
titles = []
for title in natsorted(glob.glob('imagenes/*.png')):
    titles.append(os.path.basename(title))

# insert titles
for index, title in enumerate(titles):
    worksheet.cell(row=index+2, column=7, value=title)

# save workbook
workbook.save('salida.xlsx')