import io
import json
import numpy
import re
import time
from datetime import date
from unicodedata import normalize
from webdriver_manager.chrome import ChromeDriverManager
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from PIL import Image as PImage
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as DImage
from openpyxl.utils import get_column_letter


class WebScraper (object):
    product_list: list = []
    categories_data: dict = {}
    cities_data: dict = {}

    # driver = 'D:\ProyectosCarToro\scraping\webscraper\webscraper\ChromeSetup.exe'
    chrome_options = webdriver.ChromeOptions()
    chrome_options.set_capability("acceptInsecureCerts", True)
    chrome_options.add_argument('--disable-gpu')
    chrome_options.add_argument('--mute-audio')
    chrome_options.add_argument('--ignore-certificate-errors')
    chrome_options.add_argument('--ignore-ssl-errors')
    chrome_options.add_argument('--disable-infobars')
    chrome_options.add_argument('--ignore-certificate-errors-spki-list')
    chrome_options.add_argument('--no-sandbox')
    chrome_options.add_argument('--no-zygote')
    chrome_options.add_argument('--log-level=3')
    chrome_options.add_argument('--allow-running-insecure-content')
    chrome_options.add_argument('--disable-web-security')
    chrome_options.add_argument('--disable-features=VizDisplayCompositor')
    chrome_options.add_argument('--disable-breakpad')

    desired_capabilities = chrome_options.to_capabilities()
    # chrome_options.binary_location = chrome_options.binary_location = "C:\Program Files\Google\Chrome Beta\Application\chrome.exe"
    driver = webdriver.Chrome(service=ChromeService(
        ChromeDriverManager().install()), options=chrome_options, desired_capabilities=desired_capabilities)
    # driver.minimize_window()
    driver.maximize_window()
    driver.implicitly_wait(4)

    driver.get("https://www.homecenter.com.co")
    html = driver.page_source
    time.sleep(1)
    botones: list = []
    parent_categories: list = []
    child_categories: list = []
    node_count = 0
    list_categories: list = []
    average_rate_by_product = 9

    def get_categories_params(self):
        with open('./categories.json', 'r') as f:
            data = f.read()
            f.close()
        self.categories_data = json.loads(data)

    def get_cities_params(self):
        with open('./cities.json', 'r') as f:
            data = f.read()
            f.close()
        self.cities_data = json.loads(data)

    def scan_products(self):
        print("Escaneando ... ")
        for category_i in range(len(self.list_categories)):
            self.driver.get(
                self.list_categories[category_i]["link"]+"?currentpage=1&sortBy=variant.name,asc")
            time.sleep(2)
            # se inician guardando registros primer resultado (pagina 1)
            self.list_categories[category_i]["products"] = numpy.concatenate(
                (self.list_categories[category_i]["products"], self.get_link_products()))

            totalButttonsPagination = self.get_total_buttons_by_pagination()

            # se guardan registros de la paginacion desde la pagina 2
            for point_links in range(len(totalButttonsPagination)-1):
                self.driver.get(
                    self.list_categories[category_i]["link"]+f"?currentpage={point_links+2}&sortBy=variant.name,asc")
                self.list_categories[category_i]["products"] = numpy.concatenate(
                    (self.list_categories[category_i]["products"], self.get_link_products()))

        print("Escaneado completo")
        print(self.list_categories)

    def get_link_products(self):
        list_products = []
        time.sleep(2)
        js_script = '''\
        var banner= document.getElementById('banner-plp');
        var banner2= document.getElementById('testId-input-typeahead-desktop');
        var banner3= document.getElementById('first-container-SmartAppBanner-b89b04a5-444c-448a-9f74-29fa30cf6487');

         if(banner){
            banner.setAttribute("hidden","");
        }
        if(banner2){
            banner2.setAttribute("hidden","");
        }
        if(banner3){
            banner3.setAttribute("hidden","");
        }
        '''
        self.driver.execute_script(js_script)

        grid = WebDriverWait(self.driver, 5).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="testId-btn-grid-view"]')))
        if (grid):
            grid.click()
        product_link = self.driver.find_elements(
            By.XPATH, '//*[@id="title-pdp-link"]')
        for link in product_link:
            list_products.append({'link': link.get_attribute("href"),
                                  'id': link.get_attribute("href").split("/")[-2]})

        return list_products

    def get_total_buttons_by_pagination(self):
        time.sleep(6)
        buttons: list = []
        js_script = '''\
        var banner= document.getElementById('banner-plp');
        var banner2= document.getElementById('testId-input-typeahead-desktop');
        var banner3= document.getElementById('first-container-SmartAppBanner-b89b04a5-444c-448a-9f74-29fa30cf6487');

       if(banner){
            banner.setAttribute("hidden","");
        }
        if(banner2){
            banner2.setAttribute("hidden","");
        }
        if(banner3){
            banner3.setAttribute("hidden","");
        }
        '''
        self.driver.execute_script(js_script)

        try:
            buttons = self.driver.find_element(
                By.XPATH,
                '//*[@id="__next"]/div/div/div[6]/div[3]/div[1]/div[1]/div[2]/div/div/div[2]/div/div[1]/ul').find_elements(By.CSS_SELECTOR, "button.jsx-4278284191")
            return buttons
        except:
            "No se encuentra la paginacion."
        try:
            buttons = self.driver.find_element(
                By.XPATH,
                '//*[@id="__next"]/div/div/div[7]/div[3]/div[1]/div[1]/div[2]/div/div/div[2]/div/div[1]/ul').find_elements(By.CSS_SELECTOR, "button.jsx-4278284191")
            return buttons
        except:
            "No se encuentra la paginacion, se asume no tiene."
        return buttons

    def map_product_data(self):
        workbook = load_workbook(filename='template.xlsx')
        worksheet = workbook.active
        excel_row = 0
        # resize cells
        print("TOTAL PRODUCTOS: "+str(self.get_total_products()))
        for row in range(2, self.get_total_products()+2):
            worksheet.row_dimensions[row].height = 140
            col_letter = get_column_letter(34)
            worksheet.column_dimensions[col_letter].width = 30

        for list_category_i in range(len(self.list_categories)):
            print("PRODUCTOS ESCANEADOS: "+str(excel_row) +
                  " RESTANTES: "+str(self.get_total_products()-excel_row))
            estimated_time = (
                ((self.get_total_products()-(excel_row+1))*self.average_rate_by_product))/3600
            print("TIEMPO ESTIMADO: "+str(round(estimated_time, 2))+" horas")
            for products_i in range(len(self.list_categories[list_category_i]["products"])):
                try:
                    print(f'[SCAN] PRODUCTO {str(excel_row)}: ' +
                          self.list_categories[list_category_i]["products"][products_i]['link'])
                    locations = []
                    self.driver.get(
                        self.list_categories[list_category_i]["products"][products_i]['link'])
                    time.sleep(2)
                    titulo = self._normalice_string(self.driver.find_element(
                        By.XPATH, '//*[@id="__next"]/div/div/div[4]/div[2]/div[1]/div[1]/h1').text)
                    precio = self.driver.find_element(
                        By. XPATH, '//div[@class="jsx-2167963490 primary"]/span[2]').text
                    marca = self._normalice_string(self.driver.find_element(
                        By. XPATH, '//*[@id="__next"]/div/div/div[4]/div[2]/div[1]/div[1]/div[1]/div[1]').text)
                    dataSheet = self.map_datasheet()
                    try:
                        home_delivery = self._normalice_string(self.driver.find_element(
                            By. XPATH, '//*[@id="__next"]/div/div/div[4]/div[2]/div[3]/div[2]/div[2]/div[1]/div[1]').text)
                    except:
                        home_delivery = ""
                    try:
                        pick_up_in_store = self._normalice_string(self.driver.find_element(
                            By. XPATH, '//*[@id="__next"]/div/div/div[4]/div[2]/div[3]/div[4]/div[2]/div[1]/div[1]').text)
                    except:
                        pick_up_in_store = ""
                    try:
                        stock_in_store = self._normalice_string(self.driver.find_element(
                            By. XPATH, '//*[@id="__next"]/div/div/div[4]/div[2]/div[3]/div[5]/div[2]/div[1]/div[1]').text)
                    except:
                        stock_in_store = ""
                    categories = self.list_categories[list_category_i]["categories"]
                    # find and save image
                    image_path = ".\\imagenes\\" + \
                        self.list_categories[list_category_i]["products"][products_i]['id'] + ".png"
                    image = self.driver.find_element(
                        By.XPATH, '//*[@id="pdpMainImage-' + self.list_categories[list_category_i]["products"][products_i]['id'] + '"]')
                    result = image.screenshot_as_png
                    image_to_save = PImage.open(io.BytesIO(result))
                    image_to_save.thumbnail((150, 150), PImage.LANCZOS)
                    image_to_save.save(image_path, optimize=True, quality=60)
                    time.sleep(2)

                    for index_categories in range(len(categories)):
                        worksheet.cell(row=excel_row+2, column=index_categories+1,
                                       value=categories[index_categories])

                    worksheet.cell(row=excel_row+2,
                                   column=6, value=marca)
                    worksheet.cell(row=excel_row+2,
                                   column=7, value=titulo)
                    if "modelo" in dataSheet:
                        worksheet.cell(row=excel_row+2,
                                       column=8, value=dataSheet['modelo'])
                    worksheet.cell(row=excel_row+2,
                                   column=9, value=precio)
                    if "coleccion" in dataSheet:
                        worksheet.cell(row=excel_row+2,
                                       column=10, value=dataSheet['coleccion'])
                    if "tipo" in dataSheet:
                        worksheet.cell(row=excel_row+2,
                                       column=11, value=dataSheet['tipo'])
                    if "dimensiones" in dataSheet:
                        worksheet.cell(row=excel_row+2,
                                       column=12, value=dataSheet['dimensiones'])
                    if "largo_(cm)" in dataSheet:
                        worksheet.cell(row=excel_row+2,
                                       column=13, value=dataSheet['largo_(cm)'])
                    if "largo" in dataSheet:
                        worksheet.cell(row=excel_row+2,
                                       column=13, value=dataSheet['largo'])
                    if "ancho_(cm)" in dataSheet:
                        worksheet.cell(row=excel_row+2,
                                       column=14, value=dataSheet['ancho_(cm)'])
                    if "ancho" in dataSheet:
                        worksheet.cell(row=excel_row+2,
                                       column=14, value=dataSheet['ancho'])
                    if "alto_(cm)" in dataSheet:
                        worksheet.cell(row=excel_row+2,
                                       column=15, value=dataSheet['alto_(cm)'])
                    if "alto" in dataSheet:
                        worksheet.cell(row=excel_row+2,
                                       column=15, value=dataSheet['alto'])
                    if "diametro" in dataSheet:
                        worksheet.cell(row=excel_row+2,
                                       column=16, value=dataSheet['diametro'])
                    if "peso" in dataSheet:
                        worksheet.cell(row=excel_row+2,
                                       column=17, value=dataSheet['peso'])
                    if "capacidad_volumetrica" in dataSheet:
                        worksheet.cell(row=excel_row+2,
                                       column=18, value=dataSheet['capacidad_volumetrica'])
                    if "numero_de_piezas" in dataSheet:
                        worksheet.cell(row=excel_row+2,
                                       column=19, value=dataSheet['numero_de_piezas'])

                    if "color" in dataSheet:
                        worksheet.cell(row=excel_row+2,
                                       column=20, value=dataSheet['color'])
                    if "material" in dataSheet:
                        worksheet.cell(row=excel_row+2,
                                       column=21, value=dataSheet['material'])
                    if "forma" in dataSheet:
                        worksheet.cell(row=excel_row+2,
                                       column=22, value=dataSheet['forma'])
                    if "uso_(domestico_o/y_institucional)" in dataSheet:
                        worksheet.cell(row=excel_row+2,
                                       column=23, value=dataSheet['uso_(domestico_o/y_institucional)'])
                    if "uso" in dataSheet:
                        worksheet.cell(row=excel_row+2,
                                       column=23, value=dataSheet['uso'])
                    if "origen" in dataSheet:
                        worksheet.cell(row=excel_row+2,
                                       column=24, value=dataSheet['origen'])
                    if "pais_de_origen" in dataSheet:
                        worksheet.cell(row=excel_row+2,
                                       column=24, value=dataSheet['pais_de_origen'])
                    if "procedencia" in dataSheet:
                        worksheet.cell(row=excel_row+2,
                                       column=24, value=dataSheet['procedencia'])
                    if "garantia" in dataSheet:
                        worksheet.cell(row=excel_row+2,
                                       column=25, value=dataSheet['garantia'])
                    if "caracteristicas" in dataSheet:
                        worksheet.cell(row=excel_row+2,
                                       column=26, value=dataSheet['caracteristicas'])
                    if "contenido" in dataSheet:
                        worksheet.cell(row=excel_row+2,
                                       column=27, value=dataSheet['contenido'])
                    worksheet.cell(row=excel_row+2,
                                   column=28, value=home_delivery)
                    worksheet.cell(row=excel_row+2,
                                   column=29, value=pick_up_in_store)
                    worksheet.cell(row=excel_row+2,
                                   column=30, value=stock_in_store)
                    today = date.today()
                    worksheet.cell(row=excel_row+2,
                                   column=31, value=today)
                    worksheet.cell(row=excel_row+2,
                                   column=32, value=self.list_categories[list_category_i]["products"][products_i]['link'])
                    worksheet.cell(row=excel_row+2,
                                   column=33, value=self.list_categories[list_category_i]["products"][products_i]['id'])

                    worksheet.add_image(DImage(image_path),
                                        anchor='AH'+str(excel_row+2))

                    for index_cities in range(len(self.cities_data['departments'])):
                        scan = self.scan_city(self.cities_data['departments'][index_cities]
                                              ['name'], self.cities_data['departments'][index_cities]['city'])
                        if (len(scan) > 0):
                            locations = numpy.concatenate((locations, scan))

                    for index_location in range(len(locations)):
                        if ('calle26' in locations[index_location]['city_name'].lower().replace(" ", "")):
                            worksheet.cell(row=excel_row+2,
                                           column=35, value=locations[index_location]['stock_quantity'])
                        if ('cedritos' in locations[index_location]['city_name'].lower().replace(" ", "")):
                            worksheet.cell(row=excel_row+2,
                                           column=36, value=locations[index_location]['stock_quantity'])
                        if ('av.68sur' in locations[index_location]['city_name'].lower().replace(" ", "")):
                            worksheet.cell(row=excel_row+2,
                                           column=37, value=locations[index_location]['stock_quantity'])
                        if ('tintal' in locations[index_location]['city_name'].lower().replace(" ", "")):
                            worksheet.cell(row=excel_row+2,
                                           column=38, value=locations[index_location]['stock_quantity'])
                        if ('suba' in locations[index_location]['city_name'].lower().replace(" ", "")):
                            worksheet.cell(row=excel_row+2,
                                           column=39, value=locations[index_location]['stock_quantity'])
                        if ('calima' in locations[index_location]['city_name'].lower().replace(" ", "")):
                            worksheet.cell(row=excel_row+2,
                                           column=40, value=locations[index_location]['stock_quantity'])
                        if ('calle170' in locations[index_location]['city_name'].lower().replace(" ", "")):
                            worksheet.cell(row=excel_row+2,
                                           column=41, value=locations[index_location]['stock_quantity'])
                        if ('c.c.mercurio' in locations[index_location]['city_name'].lower().replace(" ", "")):
                            worksheet.cell(row=excel_row+2,
                                           column=42, value=locations[index_location]['stock_quantity'])
                        if ('mosquera' in locations[index_location]['city_name'].lower().replace(" ", "")):
                            worksheet.cell(row=excel_row+2,
                                           column=43, value=locations[index_location]['stock_quantity'])
                        if ('cajic' in locations[index_location]['city_name'].lower().replace(" ", "")):
                            worksheet.cell(row=excel_row+2,
                                           column=44, value=locations[index_location]['stock_quantity'])
                        if ('girardot' in locations[index_location]['city_name'].lower().replace(" ", "")):
                            worksheet.cell(row=excel_row+2,
                                           column=45, value=locations[index_location]['stock_quantity'])
                        if ('calinorte' in locations[index_location]['city_name'].lower().replace(" ", "")):
                            worksheet.cell(row=excel_row+2,
                                           column=46, value=locations[index_location]['stock_quantity'])
                        if ('c.cjard�nplaza' in locations[index_location]['city_name'].lower().replace(" ", "")):
                            worksheet.cell(row=excel_row+2,
                                           column=47, value=locations[index_location]['stock_quantity'])
                        if ('palmira,unicentro' in locations[index_location]['city_name'].lower().replace(" ", "")):
                            worksheet.cell(row=excel_row+2,
                                           column=48, value=locations[index_location]['stock_quantity'])
                        if ('tulua' in locations[index_location]['city_name'].lower().replace(" ", "")):
                            worksheet.cell(row=excel_row+2,
                                           column=49, value=locations[index_location]['stock_quantity'])
                        if ('barranquillacalle30' in locations[index_location]['city_name'].lower().replace(" ", "")):
                            worksheet.cell(row=excel_row+2,
                                           column=50, value=locations[index_location]['stock_quantity'])
                        if ('barranquillanorte' in locations[index_location]['city_name'].lower().replace(" ", "")):
                            worksheet.cell(row=excel_row+2,
                                           column=51, value=locations[index_location]['stock_quantity'])
                        if ('barranquillacentro' in locations[index_location]['city_name'].lower().replace(" ", "")):
                            worksheet.cell(row=excel_row+2,
                                           column=52, value=locations[index_location]['stock_quantity'])
                        if ('cartagenalapopa' in locations[index_location]['city_name'].lower().replace(" ", "")):
                            worksheet.cell(row=excel_row+2,
                                           column=53, value=locations[index_location]['stock_quantity'])
                        if ('cartagenasanfernando' in locations[index_location]['city_name'].lower().replace(" ", "")):
                            worksheet.cell(row=excel_row+2,
                                           column=54, value=locations[index_location]['stock_quantity'])
                        if ('santamartac.c.buenavista' in locations[index_location]['city_name'].lower().replace(" ", "")):
                            worksheet.cell(row=excel_row+2,
                                           column=55, value=locations[index_location]['stock_quantity'])
                        if ('medell�nc.c.molinos' in locations[index_location]['city_name'].lower().replace(" ", "")):
                            worksheet.cell(row=excel_row+2,
                                           column=56, value=locations[index_location]['stock_quantity'])
                        if ('medell�n,sanjuan' in locations[index_location]['city_name'].lower().replace(" ", "")):
                            worksheet.cell(row=excel_row+2,
                                           column=57, value=locations[index_location]['stock_quantity'])
                        if ('medell�n,envigado' in locations[index_location]['city_name'].lower().replace(" ", "")):
                            worksheet.cell(row=excel_row+2,
                                           column=58, value=locations[index_location]['stock_quantity'])
                        if ('medell�n,industriales' in locations[index_location]['city_name'].lower().replace(" ", "")):
                            worksheet.cell(row=excel_row+2,
                                           column=59, value=locations[index_location]['stock_quantity'])
                        if ('bello' in locations[index_location]['city_name'].lower().replace(" ", "")):
                            worksheet.cell(row=excel_row+2,
                                           column=60, value=locations[index_location]['stock_quantity'])
                        if ('rionegro' in locations[index_location]['city_name'].lower().replace(" ", "")):
                            worksheet.cell(row=excel_row+2,
                                           column=61, value=locations[index_location]['stock_quantity'])

                    excel_row += 1

                except Exception as e:
                    print("[ERROR] PRODUCTO NO ENCONTRADO: " +
                          self.list_categories[list_category_i]["products"][products_i]['link'])
                    print("[ERROR]  " + repr(e))
                    continue

        print("EJECUCION COMPLETA, archivo: salida.xlsx")
        workbook.save('salida.xlsx')
        workbook.close()
        self.driver.quit()

    def get_total_products(self):
        total = 0
        for list_category_i in range(len(self.list_categories)):
            total = total + \
                len(self.list_categories[list_category_i]["products"])
        return total

    def _convert_from_array_to_object(self, arr):
        prop: dict = {}
        for x in range(0, len(arr), 2):
            prop[arr[x].lower().replace(" ", "_")] = arr[x+1]
            print(prop)
        return prop

    def _normalice_string(self, text):
        return re.sub(
            r"([^n\u0300-\u036f]|n(?!\u0303(?![\u0300-\u036f])))[\u0300-\u036f]+", r"\1",
            normalize("NFD", text), 0, re.I
        ).strip()

    def map_datasheet(self):
        dat = {}
        data_sheet = []
        try:
            data_sheet = self.driver.find_element(
                By.XPATH, '//*[@id="pdp-highlights"]').find_elements(By.CSS_SELECTOR, "div.jsx-3969330179.row")
        except:
            print("No se encontro ficha tecnica")
        for data in data_sheet:
            detail = self._normalice_string(
                data.text).split("\n")
            dat[detail[0].lower().replace(" ", "_")] = detail[1]
        return dat

    def load_data(self):
        for key in self.categories_data:
            self.element_depth(self.categories_data, key, [], True)

    def element_depth(self, grapho, current_element, analize_elements=[], reset=False):
        if reset:
            self.parent_categories = []
            self.child_categories = []
            self.node_count = 0

        if current_element in analize_elements:
            return

        analize_elements.append(current_element)
        print("elemento: {}".format(current_element))

        for neighbor in grapho[current_element]:
            if "category_" in neighbor:
                if grapho[current_element]["name"] not in self.parent_categories:
                    self.parent_categories.append(
                        grapho[current_element]["name"])

                self.element_depth(
                    grapho[current_element], neighbor, analize_elements)

            elif "end" in neighbor:
                self.node_count = self.node_count+1
                # valida si la solo hay una subcategoria y lo asocia como padre por defecto, indicando un solo nivel
                if len(self.parent_categories) == 0:
                    self.parent_categories.append(
                        grapho[current_element]["name"])
                else:
                    self.child_categories.append(
                        {"name": grapho[current_element]["name"],
                         "link": grapho[current_element]["link"]})
                # Se valida si se ha completada las ramificaciones o si la rama es de un solo nivel
                if ((self.node_count+2 == len(grapho)) or
                        (len(self.parent_categories) == 1 and self.node_count+2 == len(grapho[current_element]))):
                    print("Final de la rama")
                    if len(self.child_categories) >= 1:
                        for child in self.child_categories:
                            # copiar elementos de parent_categories en otro array
                            final_category = [None]*len(self.parent_categories)
                            for i in range(0, len(self.parent_categories)):
                                final_category[i] = self.parent_categories[i]

                            final_category.append(child["name"])
                            self.list_categories.append(
                                {"link": child["link"],
                                 "categories": final_category,
                                 "products": []})
                            final_category = []
                    else:
                        # copiar elementos de parent_categories en otro array
                        final_category = [None]*len(self.parent_categories)
                        for i in range(0, len(self.parent_categories)):
                            final_category[i] = self.parent_categories[i]

                        self.list_categories.append(
                            {"link": grapho[current_element]["link"],
                             "categories": final_category,
                             "products": []})

                        final_category = []

                    # reiniciar escaneo
                    self.parent_categories = [self.parent_categories[0]]
                    self.child_categories = []
                    self.node_count = 0

    def scan_city(self, department_name, city_name):
        final_locations = []

        try:
            time.sleep(.5)
            js_script = '''\
            var banner= document.getElementById('banner-plp');
            var banner2= document.getElementById('testId-input-typeahead-desktop');
            var banner3= document.getElementById('first-container-SmartAppBanner-b89b04a5-444c-448a-9f74-29fa30cf6487');

            if(banner){
                banner.setAttribute("hidden","");
            }
            if(banner2){
                banner2.setAttribute("hidden","");
            }
            if(banner3){
                banner3.setAttribute("hidden","");
            }
            '''
            self.driver.execute_script(js_script)
            js_script = '''\
            window.scrollTo({
            top: 0,
            left: 0,
            behavior: 'smooth'
            });
            '''
            self.driver.execute_script(js_script)
            time.sleep(2)
            WebDriverWait(self.driver, 20).until(
                EC.element_to_be_clickable((By.XPATH, '//*[@id="locationv3-content"]/div/p/span[2]'))).click()
            time.sleep(0.7)
            WebDriverWait(self.driver, 20).until(
                EC.element_to_be_clickable((By.XPATH, '//*[@id="locationv3-modal"]/div[1]/div[2]/div/div/div[3]/div[1]/div/div[2]/button'))).click()
            time.sleep(0.7)
            WebDriverWait(self.driver, 20).until(
                EC.element_to_be_clickable((By.XPATH, '//*[@id="dropdown-input-test"]'))).send_keys(department_name)
            time.sleep(0.7)
            WebDriverWait(self.driver, 20).until(
                EC.element_to_be_clickable((By.XPATH, '//*[@id="locationv3-modal"]/div[1]/div[2]/div/div/div[3]/div[1]/div/div[2]/div[2]/div[2]/button'))).click()
            time.sleep(0.7)
            WebDriverWait(self.driver, 20).until(
                EC.element_to_be_clickable((By.XPATH, '//*[@id="locationv3-modal"]/div[1]/div[2]/div/div/div[3]/div[2]/div/div[2]/button'))).click()
            time.sleep(0.7)
            WebDriverWait(self.driver, 20).until(
                EC.element_to_be_clickable((By.XPATH, '//*[@id="dropdown-input-test"]'))).send_keys(city_name)
            time.sleep(0.7)
            WebDriverWait(self.driver, 20).until(
                EC.element_to_be_clickable((By.XPATH, '//*[@id="locationv3-modal"]/div[1]/div[2]/div/div/div[3]/div[2]/div/div[2]/div[2]/div[2]/button[1]'))).click()
            time.sleep(0.7)
            WebDriverWait(self.driver, 20).until(
                EC.element_to_be_clickable((By.XPATH, '//*[@id="locationv3-modal"]/div[1]/div[2]/div/div/div[4]/button'))).click()
            time.sleep(2)
            WebDriverWait(self.driver, 20).until(
                EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/div/div/div[4]/div[2]/div[3]/div[5]/div[2]/div[2]/button'))).click()

            time.sleep(1)
            js_script = '''\
            var scrollBar = document.querySelector("div.jsx-1116510066.stores-container.jsx-150557787");
            if(scrollBar){
                scrollBar.scrollTop=10000;
            }
            '''
            self.driver.execute_script(js_script)
            time.sleep(1)

            locations = self.driver.find_elements(
                By.CLASS_NAME, 'jsx-626129325.store-details')

            for location in locations:
                detail = self._normalice_string(location.text).split("\n")
                final_locations.append({
                    'city_name': detail[0] if len(detail) > 0 else 'not-found',
                    'direction': detail[1] if len(detail) > 1 else 'not-found',
                    'stock_quantity': detail[2] if len(detail) > 2 else 'not-found',
                })

            time.sleep(0.5)
            WebDriverWait(self.driver, 20).until(
                EC.element_to_be_clickable((By.XPATH, '//*[@id="__next"]/div/div/div[4]/div[2]/div[3]/div[6]/div[1]/span'))).click()
            time.sleep(0.5)

        except Exception as e:
            print("[ERROR] ESCANENADO CIUDAD: " +
                  department_name + " - "+city_name)
            print("[ERROR]  " + repr(e))

        return final_locations


webScraper = WebScraper()
webScraper.get_categories_params()
webScraper.get_cities_params()
webScraper.load_data()
webScraper.scan_products()
webScraper.map_product_data()
